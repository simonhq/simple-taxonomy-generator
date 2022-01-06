
#to make executable run the following:
#chmod +x tgen.py
#add to path variable

#!/usr/bin/env python

from rdflib import Graph, URIRef, Literal, Namespace, BNode, plugin
from rdflib.namespace import RDF, RDFS, SKOS, OWL, PROV, XSD, SDO, DCTERMS
from tkinter import *
from tkinter import ttk, messagebox
from datetime import datetime
from openpyxl import load_workbook
from utils import *

#todays date
dtoday = datetime.today().strftime('%Y-%m-%d')
#string cleanup
REPLACEALL = str.maketrans({" ": "_", "(": "_", ")": "_", "&": "_and_", "|": "", "-": ""}) #this also replaces spaces with underscores
#file name
CSVF = "./CSTax.xlsx" #change this to the name of the csv file to consume
TAXA = "./out_tax"+ str(dtoday) +".ttl" #change this to the output file to write the taxonomy to

#variables
ontname = "cybersecurity" #change this to the name of the taxonomy
taxtype = "csdomain" #change this to the name of the schema
taxtypename  = "A Cyber Security Domain" #change this to the display name of the schema
artclass = "article" #change this the specific class that inherits from skos:Concept (or just use a keyword)
artclassname = "Article" #change this the specific class that inherits from skos:Concept (or just use a keyword)
arttop = "articles" #change this the specific name of the top level concept for the articles
conclass = "vocabulary" #change this the specific class that inherits from skos:Concept (or just use a keyword)
conclassname  = "Cyber Security Vocabulary" #change this to the display name of the concepts
contop = "vocabs" #change this the specific name of the top level concept for the keywords

#Namespaces
ONT = Namespace("https://data.uc.edu.au/def/" + ontname + "/")
TAX = URIRef("https://data.uc.edu.au/def/" + taxtype)
ART = URIRef("https://data.uc.edu.au/def/" + ontname + "/" + artclass + "#")
ELE = URIRef("https://data.uc.edu.au/def/" + ontname + "/" + conclass + "#")

ARTC = URIRef(ART + arttop)
ELEC = URIRef(ELE + contop)

#BASEURI & IMPORTURI to add after serialization (this is specific for topbraid but doesn't stop other tools working as well)
BASEURI = "# baseuri: https://data.uc.edu.au/def/" + taxtype
IMPORTURI = "# imports: https://data.uc.edu.au/def/" + ontname

#file control
artcount = 2   #this is the first row in the article worksheet that holds data
keycount = 2   #this is the first row in the keyword worksheet that holds data
active_row = 0   #this is a loop counter

def maketax(*args):
   try:
      #print(mfile.get())

      #---------------------------------------------------------------------
      # create a new graph
      atax = Graph()
      # bind the prefixes
      bind_prefixes(atax)
      atax.bind("ONT", ONT)
      
      #atax.add((TAX, RDF.type, OWL.Ontology))
      #atax.add((TAX, RDFS.label, Literal(taxtypename)))

      #create the conceptschmeme and top level 
      atax.add((TAX, RDF.type, SKOS.ConceptScheme))
      atax.add((TAX, SKOS.prefLabel, Literal(taxtypename)))

      #articles top level
      atax.add((ARTC, RDF.type, SKOS.Concept))
      atax.add((TAX, SKOS.hasTopConcept, ARTC))
      atax.add((ARTC, SKOS.prefLabel, Literal(artclassname)))

      #keywords top level
      atax.add((ELEC, RDF.type, SKOS.Concept))
      atax.add((TAX, SKOS.hasTopConcept, ELEC))
      atax.add((ELEC, SKOS.prefLabel, Literal(conclassname)))


      #---------------------------------------------------------------------

      wb = load_workbook(filename=mfile.get())
      ws = wb['Article']

      active_row = 0
      for row in ws.rows:
         active_row += 1
         if active_row < artcount:
            #this is the instructions and headings
            pass
         else:
            vid = str(ws.cell(active_row,column=1).value).strip().translate(REPLACEALL)
            vtitle = ws.cell(active_row,column=2).value
            vyear = ws.cell(active_row,column=3).value
            vauth = ws.cell(active_row,column=4).value
            vrefset = str(ws.cell(active_row,column=5).value).strip().split("|")
            
            atax.add((URIRef(ART + vid), RDF.type, SKOS.ConceptScheme))
            atax.add((URIRef(ART + vid), SKOS.prefLabel, Literal(vtitle)))
            atax.add((URIRef(ART + vid), RDF.type, DCTERMS.BibliographicResource))
            atax.add((URIRef(ART + vid), DCTERMS.date, Literal(vyear)))
            atax.add((URIRef(ART + vid), DCTERMS.source, Literal(vtitle)))
            atax.add((URIRef(ART + vid), DCTERMS.creator, Literal(vauth)))

            for vref in vrefset:
               if vref != 'None':
                  atax.add((URIRef(ART + vid), DCTERMS.references, URIRef(ART + vref)))
      
            #####################################################################################################
            #now add it to the master list
            vit = str(vtitle).strip().translate(REPLACEALL)
            atax.add((URIRef(ART + vit), RDF.type, SKOS.Concept))
            atax.add((URIRef(ART + vit), SKOS.prefLabel, Literal(vtitle)))
            atax.add((URIRef(ART + vit), SKOS.exactMatch, URIRef(ART + vid)))
            atax.add((URIRef(ART + vit), SKOS.broader, ARTC))
            #####################################################################################################

      ws = wb['Keyword']

      active_row = 0
      for row in ws.rows:
         active_row += 1
         if active_row < keycount:
            #this is the instructions and headings
            pass
         else:

            #read and build the term detail
            eval = ""
            lval = ""
            tval = str(ws.cell(active_row,column=3).value).strip().split("|")
            tcnt = len(tval) #get the number of terms
            #get the main link
            for i in range(0,tcnt):
               lval = eval
               eval = eval + tval[i].strip().translate(REPLACEALL)
               vid = str(ws.cell(active_row,column=2).value).strip().translate(REPLACEALL)

               if tval[i] != 'None':

                  atax.add((URIRef(ELE + vid + eval), RDF.type, SKOS.Concept))
                  atax.add((URIRef(ELE + vid + eval), SKOS.prefLabel, Literal(tval[i].strip().capitalize())))
                  #link them all to the article 
                  atax.add((URIRef(ELE + vid + eval), SKOS.inScheme, URIRef(ART + vid)))
                  #atax.add((URIRef(ELE + eval), DCTERMS.bibliographicCitation, URIRef(ART + vid)))
                  if i == 0:
                     atax.add((URIRef(ART + vid), SKOS.hasTopConcept, URIRef(ELE + vid + eval)))
                     atax.add((URIRef(ELE + vid + eval), SKOS.broader, URIRef(ART + vid)))
                  else:
                     atax.add((URIRef(ELE + vid + eval), SKOS.broader, URIRef(ELE + vid + lval)))

                  #####################################################################################################
                  #now add it to the master list
                  if i == 0:
                     atax.add((URIRef(ELE + eval), RDF.type, SKOS.Concept))
                     atax.add((URIRef(ELE + eval), SKOS.prefLabel, Literal(tval[i].strip().capitalize())))
                     atax.add((URIRef(ELE + eval), SKOS.broader, ELEC))
                     atax.add((URIRef(ELE + vid + eval), SKOS.closeMatch, URIRef(ELE + eval)))
                  else:
                     atax.add((URIRef(ELE + eval), RDF.type, SKOS.Concept))
                     atax.add((URIRef(ELE + eval), SKOS.prefLabel, Literal(tval[i].strip().capitalize())))
                     atax.add((URIRef(ELE + eval), SKOS.broader, URIRef(ELE + lval)))
                     atax.add((URIRef(ELE + vid + eval), SKOS.closeMatch, URIRef(ELE + eval)))
                  #####################################################################################################

            #now that the eval is the final link - add definition
            kdef = ws.cell(active_row,column=5).value
            if kdef != None:
               atax.add((URIRef(ELE + vid + eval), SKOS.definition, Literal(str(kdef).strip().capitalize())))
            
            #####################################################################################################
            #add the link to a direct reference
            rdef = ws.cell(active_row,column=6).value
            if rdef != None:
               rdef = str(rdef).strip()
               atax.add((URIRef(ELE + rdef + eval), RDF.type, SKOS.Concept))
               atax.add((URIRef(ELE + rdef + eval), SKOS.inScheme, URIRef(ART + rdef)))
               atax.add((URIRef(ELE + rdef + eval), SKOS.prefLabel, Literal(ws.cell(active_row,column=3).value)))
               atax.add((URIRef(ELE + vid + eval), SKOS.exactMatch, URIRef(ELE + rdef + eval)))
               if kdef != None:
                  atax.add((URIRef(ELE + rdef + eval), SKOS.definition, Literal(str(kdef).strip().capitalize())))
            #####################################################################################################

            #now that the eval is the final link - add synonyms 
            #read and build the term detail
            esal = ""
            lsal = ""
            tsal = str(ws.cell(active_row,column=4).value).strip().split("|")
            tsnt = len(tsal) #get the number of terms
            #get the main link
            for i in range(0,tsnt):
               lsal = esal
               esal = esal + tsal[i].strip().translate(REPLACEALL)

               if tsal[i] != 'None':
                  atax.add((URIRef(ELE + vid + esal), RDF.type, SKOS.Concept))
                  atax.add((URIRef(ELE + vid + esal), SKOS.prefLabel, Literal(tsal[i].strip().capitalize())))
                  #link them all to the article
                  atax.add((URIRef(ELE + vid + esal), SKOS.inScheme, URIRef(ART + vid)))
                  if i == 0:
                     atax.add((URIRef(ART + vid), SKOS.hasTopConcept, URIRef(ELE + vid + esal)))
                  else:
                     atax.add((URIRef(ELE + vid + esal), SKOS.broader, URIRef(ELE + vid + lsal)))

                  #####################################################################################################
                  #now add it to the master list
                  if i == 0:
                     atax.add((URIRef(ELE + esal), RDF.type, SKOS.Concept))
                     atax.add((URIRef(ELE + esal), SKOS.broader, ELEC))
                     atax.add((URIRef(ELE + vid + esal), SKOS.closeMatch, URIRef(ELE + esal)))
                  else:
                     atax.add((URIRef(ELE + esal), RDF.type, SKOS.Concept))
                     atax.add((URIRef(ELE + esal), SKOS.broader, URIRef(ELE + lsal)))
                     atax.add((URIRef(ELE + vid + esal), SKOS.closeMatch, URIRef(ELE + esal)))
                  #####################################################################################################

            #link the final synonym pass to the final term pass
            if tsal[0] != 'None':
               atax.add((URIRef(ELE + vid + eval), SKOS.exactMatch, URIRef(ELE + vid + esal)))

            
      # store it in a turtle file
      atax.serialize(TAXA, format="turtle")
   
      # open and add the base URI to the start of the file (from utils.py)
      add_pragmas_to_file(TAXA, BASEURI ) # + "\n" + IMPORTURI)

   except ValueError:
      pass

   messagebox.showinfo(message=TAXA + ' File Created',title='Taxonomy builder')
   
##
# Main window code
##

mainw = Tk()
mainw.title('Taxonomy Generator')

mframe = ttk.Frame(mainw, padding="3 3 12 12")
mframe.grid(column=0, row=0, sticky=(N, W, E, S))
mainw.columnconfigure(0, weight=1)
mainw.rowconfigure(0, weight=1)

#label
ttk.Label(mframe, text='Select the file with the mappings:').grid(column=2,row=1,sticky=E)
#Now the file name
mfile = StringVar() #create a string variable
mfile.set(CSVF) #set it to the default
mfile_entry = ttk.Entry(mframe, width=25, textvariable=mfile) #put it on the page
mfile_entry.grid(column=2,row=2,sticky=(W,E)) 
#add a label for the file name
ttk.Label(mframe, text="File name").grid(column=1,row=2,sticky=E)

# run the get data
runbutt = ttk.Button(mframe, text='Run', width=25, command=maketax).grid(column=2, row=4, sticky=(W,E))

#exit button
exitbutt = ttk.Button(mframe, text='Exit', width=25, command=mainw.destroy).grid(column=3, row=4, sticky=(W,E))

#set the padding for the whole page
for child in mframe.winfo_children():
   child.grid_configure(padx=5, pady=5)

#place the focus as the text box entry
mfile_entry.focus()
#bind the return key to the run button
mainw.bind("<Return>", maketax)
#run the application
mainw.mainloop()